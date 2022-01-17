from tkinter import *
from openpyxl import *
from tkinter import ttk
from tkcalendar import DateEntry
from tkinter import messagebox
import pandas as pd
import xlrd

root = Tk()
root.title('Pujara Enterprise')
root.configure(bg='black')

list1=[]
list2=[]
list3=[]
list4=[]




date=StringVar()
datet=StringVar()
pname=StringVar()
grade=StringVar()
tname=StringVar()
dno=StringVar()
ddf=StringVar()
ddt=StringVar()
db=StringVar()
tweight=StringVar()
tno=StringVar()
totalweight=StringVar()
ulweight=StringVar()
delito=StringVar()
cp=StringVar()
sp=StringVar()
freight=StringVar()
remarks=StringVar()



wb = load_workbook('C:\\Pujara enterprise\\pujara enterprise.xlsx')

"""wb1= xlrd.open_workbook('C:\\Pujara enterprise\\Data.xlsx')"""

sheet = wb.active

"""sheet1=wb1.sheet_by_index(0)"""

def excel():


    sheet.cell(row=1, column=1).value = "Date"
    sheet.cell(row=1, column=2).value = "Grade"
    sheet.cell(row=1, column=3).value = "Party Name"
    sheet.cell(row=1, column=4).value = "Transporter Name"
    sheet.cell(row=1, column=5).value = "DO Number"
    sheet.cell(row=1, column=6).value = "DO Date From"
    sheet.cell(row=1, column=7).value = "DO Date TO"
    sheet.cell(row=1, column=8).value = "DO By"
    sheet.cell(row=1, column=9).value = "Delivered To"
    sheet.cell(row=1, column=10).value = "Truck Number"
    sheet.cell(row=1, column=11).value = "Truck Weight"
    sheet.cell(row=1, column=12).value = "Total Weight"
    sheet.cell(row=1, column=13).value = "Unloading weight"
    sheet.cell(row=1, column=14).value = "Cost Price"
    sheet.cell(row=1, column=15).value = "Sell Price"
    sheet.cell(row=1, column=16).value = "Freight"
    sheet.cell(row=1, column=17).value = "Remarks"




class app:
    def __init__(self, master):
        self.master = master
        self.master.geometry("1920x1080")
        self.maintab()

    def maintab(self):
        for i in self.master.winfo_children():
            i.destroy()
        self.title1 = Label(root, text="Welcome", bg='black', fg='White', font=('times new roman', 35, 'bold'),
                       justify=CENTER)
        self.title1.pack()

        self.title2 = Label(root, text="Pujara Enterprise Pvt. Ltd.", bg='black', fg='White',
                       font=('times new roman', 47, 'bold'))
        self.title2.place(relx=0.5, rely=0.3, anchor='center')

        self.button1 = Button(root, text='Entry tab', bg='white', fg='black', font=('times new roman', 10),
                         command=self.entrytab)
        self.button1.place(relx=0.3, rely=0.4, anchor='center')

        self.button2 = Button(root, text='Search tab', bg='white', fg='black', font=('times new roman', 10),
                         command=self.resulttab)
        self.button2.place(relx=0.5, rely=0.4, anchor='center')

        self.button3 = Button(root, text='Add', bg='white', fg='black', font=('times new roman', 10),
                         command=self.addtab)
        self.button3.place(relx=0.7, rely=0.4, anchor='center')


    def entrytab(self):
        for i in self.master.winfo_children():
            i.destroy()
        self.title1= Label(root, text="Pujara Enterprise Pvt. Ltd.", bg='white', fg='black',
                       font=('times new roman', 47, 'bold'),justify=CENTER)
        self.title1.pack(fill=X)
        self.title2 = Label(root, text="Entry Tab", bg='black', fg='White',
                            font=('times new roman', 35, 'bold'))
        self.title2.place(relx=0.5, rely=0.2, anchor='center')
        self.button1 = Button(root, text='Home', bg='white', fg='black', font=('times new roman', 10),
                              command=self.maintab)
        self.button1.place(relx=0.3, rely=0.15, anchor='center')

        self.button2 = Button(root, text='Search tab', bg='white', fg='black', font=('times new roman', 10),
                              command=self.resulttab)
        self.button2.place(relx=0.5, rely=0.15, anchor='center')

        self.button3 = Button(root, text='Add', bg='white', fg='black', font=('times new roman', 10),
                         command=self.addtab)
        self.button3.place(relx=0.7, rely=0.15, anchor='center')


        F1=LabelFrame(root, bd=10, relief=GROOVE, text='Enter details here', font=('times new romon', 15, 'bold'),
                        fg='Orange', bg='white')
        F1.place(relx=0.5, rely=0.6,anchor='center', width=1300, height=600)
        date_lbl = Label(F1, text='Date:-', font=('times new romon', 13, 'bold'), bg='White',
                          fg='#00FF33')
        date_lbl.place(relx=0.07,rely=0.07,anchor='center')
        cal = DateEntry( F1,width=15,background='darkblue',textvariable=date,date_pattern='dd/mm/yy', foreground='white', borderwidth=2)
        cal.place(relx=0.3,rely=0.07,anchor='center')



        grade_lbl = Label(F1, text='Grade:-', font=('times new romon', 13, 'bold'), bg='White',
                         fg='#00FF33')
        grade_lbl.place(relx=0.7, rely=0.07, anchor='center')
        grade_txt = Entry(F1, width=15, textvariable=grade, font='arial 10 bold', relief=SUNKEN, bd=3,
                         bg='#FFFF66')
        grade_txt.place(relx=0.87, rely=0.07, anchor='center')


        pname_lbl = Label(F1, text='Party Name:-', font=('times new romon', 13, 'bold'), bg='White',
                          fg='#00FF33')
        pname_lbl.place(relx=0.0925, rely=0.14, anchor='center')
        pname_txt = Entry(F1, width=15, textvariable=pname, font='arial 10 bold', relief=SUNKEN, bd=3,
                          bg='#FFFF66')
        pname_txt.place(relx=0.3, rely=0.14, anchor='center')


        tname_lbl = Label(F1, text='Transporter Name:-', font=('times new romon', 13, 'bold'), bg='White',
                          fg='#00FF33')
        tname_lbl.place(relx=0.112, rely=0.21, anchor='center')
        tname_txt = Entry(F1, width=15, textvariable=tname, font='arial 10 bold', relief=SUNKEN, bd=3,
                          bg='#FFFF66')
        tname_txt.place(relx=0.3, rely=0.21, anchor='center')


        dno_lbl = Label(F1, text='DO Number:-', font=('times new romon', 13, 'bold'), bg='White',
                          fg='#00FF33')
        dno_lbl.place(relx=0.093, rely=0.28, anchor='center')
        dno_txt = Entry(F1, width=15, textvariable=dno, font='arial 10 bold', relief=SUNKEN, bd=3,
                          bg='#FFFF66')
        dno_txt.place(relx=0.3, rely=0.28, anchor='center')


        ddf_lbl = Label(F1, text='DO Date from:-', font=('times new romon', 13, 'bold'), bg='White',
                          fg='#00FF33')
        ddf_lbl.place(relx=0.098, rely=0.35, anchor='center')
        ddf_txt = Entry(F1, width=15, textvariable=ddf, font='arial 10 bold', relief=SUNKEN, bd=3,
                          bg='#FFFF66')
        ddf_txt.place(relx=0.3, rely=0.35, anchor='center')


        ddt_lbl = Label(F1, text='DO Date To:-', font=('times new romon',13, 'bold'), bg='White',
                          fg='#00FF33')
        ddt_lbl.place(relx=0.718, rely=0.35, anchor='center')
        ddt_txt = Entry(F1, width=15, textvariable=ddt, font='arial 10 bold', relief=SUNKEN, bd=3,
                          bg='#FFFF66')
        ddt_txt.place(relx=0.87, rely=0.35, anchor='center')


        db_lbl = Label(F1, text='DO By:-', font=('times new romon', 13, 'bold'), bg='White',
                          fg='#00FF33')
        db_lbl.place(relx=0.07555, rely=0.42, anchor='center')
        db_txt = Entry(F1, width=15, textvariable=db, font='arial 10 bold', relief=SUNKEN, bd=3,
                          bg='#FFFF66')
        db_txt.place(relx=0.3,  rely=0.42, anchor='center')


        delito_lbl = Label(F1, text='Delivered To:-', font=('times new romon', 13, 'bold'), bg='White',
                          fg='#00FF33')
        delito_lbl.place(relx=0.72, rely=0.42, anchor='center')
        delito_txt = Entry(F1, width=15, textvariable=delito, font='arial 10 bold', relief=SUNKEN, bd=3,
                          bg='#FFFF66')
        delito_txt.place(relx=0.87, rely=0.42, anchor='center')


        tno_lbl = Label(F1, text='Truck Number:-', font=('times new romon', 13, 'bold'), bg='White',
                          fg='#00FF33')
        tno_lbl.place(relx=0.1, rely=0.49, anchor='center')
        tno_txt = Entry(F1, width=15, textvariable=tno, font='arial 10 bold', relief=SUNKEN, bd=3,
                          bg='#FFFF66')
        tno_txt.place(relx=0.3, rely=0.49, anchor='center')


        tweight_lbl = Label(F1, text='Truck Weight:-', font=('times new romon', 13, 'bold'), bg='White',
                          fg='#00FF33')
        tweight_lbl.place(relx=0.725, rely=0.49, anchor='center')
        tweight_txt = Entry(F1, width=15, textvariable=tweight, font='arial 10 bold', relief=SUNKEN, bd=3,
                          bg='#FFFF66')
        tweight_txt.place(relx=0.87, rely=0.49, anchor='center')


        totalweight_lbl = Label(F1, text='Total Weight:-', font=('times new romon', 13, 'bold'), bg='White',
                          fg='#00FF33')
        totalweight_lbl.place(relx=0.095, rely=0.56, anchor='center')
        totalweight_txt = Entry(F1, width=15, textvariable=totalweight, font='arial 10 bold', relief=SUNKEN, bd=3,
                          bg='#FFFF66')
        totalweight_txt.place(relx=0.3, rely=0.56, anchor='center')


        ulweight_lbl = Label(F1, text='Unloading Weight:-', font=('times new romon', 13, 'bold'), bg='White',
                          fg='#00FF33')
        ulweight_lbl.place(relx=0.74, rely=0.56, anchor='center')
        ulweight_txt = Entry(F1, width=15, textvariable=ulweight, font='arial 10 bold', relief=SUNKEN, bd=3,
                          bg='#FFFF66')
        ulweight_txt.place(relx=0.87, rely=0.56, anchor='center')


        cp_lbl = Label(F1, text='Cost Price:-', font=('times new romon', 13, 'bold'), bg='White',
                            fg='#00FF33')
        cp_lbl.place(relx=0.09, rely=0.63, anchor='center')
        cp_txt = Entry(F1, width=15, textvariable=cp, font='arial 10 bold', relief=SUNKEN, bd=3,
                            bg='#FFFF66')
        cp_txt.place(relx=0.3, rely=0.63, anchor='center')


        sp_lbl = Label(F1, text='Sell Price:-', font=('times new romon', 13, 'bold'), bg='White',
                            fg='#00FF33')
        sp_lbl.place(relx=0.71, rely=0.63, anchor='center')
        sp_txt = Entry(F1, width=15, textvariable=sp, font='arial 10 bold', relief=SUNKEN, bd=3,
                            bg='#FFFF66')
        sp_txt.place(relx=0.87, rely=0.63, anchor='center')


        freight_lbl = Label(F1, text='Freight:-', font=('times new romon', 13, 'bold'), bg='White',
                          fg='#00FF33')
        freight_lbl.place(relx=0.08, rely=0.70, anchor='center')
        freight_txt = Entry(F1, width=15, textvariable=freight, font='arial 10 bold', relief=SUNKEN, bd=3,
                          bg='#FFFF66')
        freight_txt.place(relx=0.3, rely=0.70, anchor='center')


        remarks_lbl = Label(F1, text='Remarks:-', font=('times new romon', 13, 'bold'), bg='White',
                          fg='#00FF33')
        remarks_lbl.place(relx=0.085, rely=0.77, anchor='center')
        remarks_txt = Entry(F1, width=15, textvariable=remarks, font='arial 10 bold', relief=SUNKEN, bd=3,
                          bg='#FFFF66')
        remarks_txt.place(relx=0.3, rely=0.77, anchor='center')

        btn1 = Button(F1, text='Clear', font='arial 15 bold', command=clear, bg='#00FFFF', width=15)
        btn1.place(relx=0.3,rely=0.93,anchor='center')

        btn2 = Button(F1, text='Submit', font='arial 15 bold', command=insert, bg='#00FFFF', width=15)
        btn2.place(relx=0.7, rely=0.93, anchor='center')


    def resulttab(self):
        for i in self.master.winfo_children():
            i.destroy()
        self.title1 = Label(root, text="Pujara Enterprise Pvt. Ltd.", bg='white', fg='black',
                            font=('times new roman', 47, 'bold'), justify=CENTER)
        self.title1.pack(fill=X)
        self.title2 = Label(root, text="Search Tab", bg='black', fg='White',
                            font=('times new roman', 35, 'bold'))
        self.title2.place(relx=0.5, rely=0.2, anchor='center')
        self.button1 = Button(root, text='Home', bg='white', fg='black', font=('times new roman', 10),
                              command=self.maintab)
        self.button1.place(relx=0.3, rely=0.15, anchor='center')

        self.button2 = Button(root, text='Entry tab', bg='white', fg='black', font=('times new roman', 10),
                              command=self.entrytab)
        self.button2.place(relx=0.5, rely=0.15, anchor='center')

        self.button3 = Button(root, text='Add', bg='white', fg='black', font=('times new roman', 10),
                         command=self.addtab)
        self.button3.place(relx=0.7, rely=0.15, anchor='center')

        F1 = LabelFrame(root, bd=10, relief=GROOVE, text='Enter details to search here', font=('times new romon', 15, 'bold'),
                        fg='Orange', bg='white')
        F1.place(relx=0.5, rely=0.6, anchor='center', width=1300, height=600)


        date_lbl = Label(F1, text='Date From:-', font=('times new romon', 13, 'bold'), bg='White',
                         fg='#00FF33')
        date_lbl.place(relx=0.09, rely=0.07, anchor='center')
        cal = DateEntry(F1, width=15, background='darkblue', textvariable=date, date_pattern='dd/mm/yy',
                        foreground='white', borderwidth=2)
        cal.place(relx=0.3, rely=0.07, anchor='center')


        date_lbl = Label(F1, text='Date To:-', font=('times new romon', 13, 'bold'), bg='White',
                         fg='#00FF33')
        date_lbl.place(relx=0.7, rely=0.07, anchor='center')
        cal = DateEntry(F1, width=15, background='darkblue', textvariable=datet, date_pattern='dd/mm/yy',
                        foreground='white', borderwidth=2)
        cal.place(relx=0.8, rely=0.07, anchor='center')

        grade_lbl = Label(F1, text='Grade:-', font=('times new romon', 13, 'bold'), bg='White',
                          fg='#00FF33')
        grade_lbl.place(relx=0.075, rely=0.14, anchor='center')
        grade_txt = Entry(F1, width=15, textvariable=grade, font='arial 10 bold', relief=SUNKEN, bd=3,
                          bg='#FFFF66')
        grade_txt.place(relx=0.3, rely=0.14, anchor='center')


        pname_lbl = Label(F1, text='Party Name:-', font=('times new romon', 13, 'bold'), bg='White',
                          fg='#00FF33')
        pname_lbl.place(relx=0.0925, rely=0.21, anchor='center')
        pname_txt = Entry(F1, width=15, textvariable=pname, font='arial 10 bold', relief=SUNKEN, bd=3,
                          bg='#FFFF66')
        pname_txt.place(relx=0.3, rely=0.21, anchor='center')


        tname_lbl = Label(F1, text='Transporter Name:-', font=('times new romon', 13, 'bold'), bg='White',
                          fg='#00FF33')
        tname_lbl.place(relx=0.112, rely=0.28, anchor='center')
        tname_txt = Entry(F1, width=15, textvariable=tname, font='arial 10 bold', relief=SUNKEN, bd=3,
                          bg='#FFFF66')
        tname_txt.place(relx=0.3, rely=0.28, anchor='center')


        delito_lbl = Label(F1, text='Delivered To:-', font=('times new romon', 13, 'bold'), bg='White',
                           fg='#00FF33')
        delito_lbl.place(relx=0.0955, rely=0.35, anchor='center')
        delito_txt = Entry(F1, width=15, textvariable=delito, font='arial 10 bold', relief=SUNKEN, bd=3,
                           bg='#FFFF66')
        delito_txt.place(relx=0.3, rely=0.35, anchor='center')


        btn1 = Button(F1, text='Clear', font='arial 15 bold', command=clear, bg='#00FFFF', width=15)
        btn1.place(relx=0.3, rely=0.85, anchor='center')

        btn2 = Button(F1, text='Submit', font='arial 15 bold', command=clear, bg='#00FFFF', width=15)
        btn2.place(relx=0.7, rely=0.85, anchor='center')


    def addtab(self):
        for i in self.master.winfo_children():
            i.destroy()
        self.title1 = Label(root, text="Pujara Enterprise Pvt. Ltd.", bg='white', fg='black',
                            font=('times new roman', 47, 'bold'), justify=CENTER)
        self.title1.pack(fill=X)
        self.title2 = Label(root, text="Search Tab", bg='black', fg='White',
                            font=('times new roman', 35, 'bold'))
        self.title2.place(relx=0.5, rely=0.2, anchor='center')
        self.button1 = Button(root, text='Home', bg='white', fg='black', font=('times new roman', 10),
                              command=self.maintab)
        self.button1.place(relx=0.3, rely=0.15, anchor='center')

        self.button2 = Button(root, text='Entry tab', bg='white', fg='black', font=('times new roman', 10),
                              command=self.entrytab)
        self.button2.place(relx=0.5, rely=0.15, anchor='center')

        self.button3 = Button(root, text='Search Tab', bg='white', fg='black', font=('times new roman', 10),
                         command=self.resulttab)
        self.button3.place(relx=0.7, rely=0.15, anchor='center')

        F1 = LabelFrame(root, bd=10, relief=GROOVE, text='Enter details to Add here', font=('times new romon', 15, 'bold'),
                        fg='Orange', bg='white')
        F1.place(relx=0.5, rely=0.6, anchor='center', width=1300, height=600)


        grade_lbl = Label(F1, text='Grade:-', font=('times new romon', 13, 'bold'), bg='White',
                          fg='#00FF33')
        grade_lbl.place(relx=0.075, rely=0.14, anchor='center')
        grade_txt = Entry(F1, width=15, textvariable=grade, font='arial 10 bold', relief=SUNKEN, bd=3,
                          bg='#FFFF66')
        grade_txt.place(relx=0.3, rely=0.14, anchor='center')


        pname_lbl = Label(F1, text='Party Name:-', font=('times new romon', 13, 'bold'), bg='White',
                          fg='#00FF33')
        pname_lbl.place(relx=0.0925, rely=0.21, anchor='center')
        pname_txt = Entry(F1, width=15, textvariable=pname, font='arial 10 bold', relief=SUNKEN, bd=3,
                          bg='#FFFF66')
        pname_txt.place(relx=0.3, rely=0.21, anchor='center')


        tname_lbl = Label(F1, text='Transporter Name:-', font=('times new romon', 13, 'bold'), bg='White',
                          fg='#00FF33')
        tname_lbl.place(relx=0.112, rely=0.28, anchor='center')
        tname_txt = Entry(F1, width=15, textvariable=tname, font='arial 10 bold', relief=SUNKEN, bd=3,
                          bg='#FFFF66')
        tname_txt.place(relx=0.3, rely=0.28, anchor='center')


        delito_lbl = Label(F1, text='Delivered To:-', font=('times new romon', 13, 'bold'), bg='White',
                           fg='#00FF33')
        delito_lbl.place(relx=0.0955, rely=0.35, anchor='center')
        delito_txt = Entry(F1, width=15, textvariable=delito, font='arial 10 bold', relief=SUNKEN, bd=3,
                           bg='#FFFF66')
        delito_txt.place(relx=0.3, rely=0.35, anchor='center')

        btn1 = Button(F1, text='Clear', font='arial 15 bold', command=clear, bg='#00FFFF', width=15)
        btn1.place(relx=0.3, rely=0.85, anchor='center')

        btn2 = Button(F1, text='Submit', font='arial 15 bold', command=dataadd, bg='#00FFFF', width=15)
        btn2.place(relx=0.7, rely=0.85, anchor='center')




def insert():

    global date, pname, grade, tname, dno, ddf, ddt, db, tweight, tno, totalweight, ulweight, delito, cp, sp, remarks, freight
    excel()

    if (    pname.get() == "" and
            grade.get()== "" and
            tname.get()== "" and
            dno.get()== "" and
            ddf.get()== "" and
            ddt.get()== "" and
            db.get()== "" and
            tweight.get()== "" and
            tno.get() == "" and
            totalweight.get()== "" and
            ulweight.get()== "" and
            delito.get()== "" and
            cp.get()== "" and
            sp.get()== "" and
            freight.get()== "" and
            remarks.get()== ""):

        messagebox.showerror("error","Please enter details!!!")
    else:
        current_row = sheet.max_row
        current_column = sheet.max_column

        sheet.cell(row=current_row + 1, column=1).value = date.get()
        sheet.cell(row=current_row + 1, column=2).value = grade.get()
        sheet.cell(row=current_row + 1, column=3).value = pname.get()
        sheet.cell(row=current_row + 1, column=4).value = tname.get()
        sheet.cell(row=current_row + 1, column=5).value = dno.get()
        sheet.cell(row=current_row + 1, column=6).value = ddf.get()
        sheet.cell(row=current_row + 1, column=7).value = ddt.get()
        sheet.cell(row=current_row + 1, column=8).value = db.get()
        sheet.cell(row=current_row + 1, column=9).value = delito.get()
        sheet.cell(row=current_row + 1, column=10).value = tno.get()
        sheet.cell(row=current_row + 1, column=11).value = tweight.get()
        sheet.cell(row=current_row + 1, column=12).value = totalweight.get()
        sheet.cell(row=current_row + 1, column=13).value = ulweight.get()
        sheet.cell(row=current_row + 1, column=14).value = cp.get()
        sheet.cell(row=current_row + 1, column=15).value = sp.get()
        sheet.cell(row=current_row + 1, column=16).value = freight.get()
        sheet.cell(row=current_row + 1, column=17).value = remarks.get()

        clear()

        date = Entry(root)
        grade = Entry(root)
        pname = Entry(root)
        tname = Entry(root)
        dno = Entry(root)
        ddf = Entry(root)
        ddt = Entry(root)
        db = Entry(root)
        delito = Entry(root)
        tno = Entry(root)
        tweight = Entry(root)
        totalweight = Entry(root)
        ulweight = Entry(root)
        cp = Entry(root)
        sp = Entry(root)
        freight = Entry(root)
        remarks = Entry(root)


        wb.save('C:\\Pujara enterprise\\pujara enterprise.xlsx')

        date.focus_set()

        excel()

def dataadd():
  """  y=0
    if pname.get() != "":
        for i in range(sheet1.nrows):
            y=y+1
        
        
        sheet1.cell(row=y + 1, column=1).value = pname.get()

    
    elif grade.get() !="":
        for i in range(sheet1.nrows):
            y=y+1
    
        sheet1.cell(row=y + 1, column=2).value = grade.get()
    
    elif tname.get() !="":
        for i in range(sheet1.nrows):
            y=y+1 
        
        sheet1.cell(row=y+ 1, column=3).value = tname.get()

    elif delito.get() !="":
        for i in range(sheet1.nrows):
            y=y+1
        
        sheet1.cell(row=y + 1, column=4).value = delito.get()

    else:
        messagebox.showerror("error","Please enter details!!!")

    wb1.save('C:\\Pujara enterprise\\Data.xlsx')
"""


def clear():
    pname.set('')
    grade.set('')
    tname.set('')
    dno.set('')
    ddf.set('')
    ddt.set('')
    db.set('')
    tweight.set('')
    tno.set('')
    totalweight.set('')                                                            
    ulweight.set('')
    delito.set('')
    cp.set('')
    sp.set('')
    freight.set('')
    remarks.set('')

app(root)
root.mainloop()